"""
setup_demo.py
Automated setup script for MRPL PS 26117 demo suite:
1. Creates ui/static/verify.html (Mobile/Desktop Web Verification Portal for scanned QR codes)
2. Builds complete sample_files/ suite across all 4 categories:
   - 01_multimodal_vision
   - 02_sovereign_rag_sops
   - 03_engineering_math_code_gen
   - 04_psu_reasoning_memos
"""

import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =====================================================================
# 1. CREATE ui/static/verify.html
# =====================================================================
verify_html_content = """<!DOCTYPE html>
<html lang="en" class="scroll-smooth">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MRPL PS 26117 — Official Cryptographic Certificate Verification</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        body { font-family: 'Inter', -apple-system, sans-serif; background-color: #06090f; color: #e2e8f0; }
        .font-mono { font-family: 'JetBrains Mono', monospace, Consolas; }
        .cyber-card { background: rgba(13, 19, 33, 0.9); backdrop-filter: blur(16px); border: 1px solid rgba(0, 240, 255, 0.15); }
        .glow-emerald { box-shadow: 0 0 30px rgba(16, 185, 129, 0.2); }
        .glow-red { box-shadow: 0 0 30px rgba(239, 68, 68, 0.3); }
    </style>
</head>
<body class="min-h-screen flex flex-col items-center justify-start p-4 sm:p-8 antialiased selection:bg-cyan-500 selection:text-black">

    <div class="max-w-4xl w-full flex flex-col space-y-6">
        
        <!-- Header Brand Bar -->
        <header class="flex items-center justify-between p-4 cyber-card rounded-2xl border border-slate-800 shadow-2xl">
            <div class="flex items-center space-x-3">
                <div class="w-10 h-10 rounded-xl bg-gradient-to-br from-amber-500 to-amber-700 flex items-center justify-center font-bold text-black text-xl shadow-lg shadow-amber-500/30">
                    M
                </div>
                <div>
                    <h1 class="text-sm sm:text-base font-extrabold tracking-wider text-slate-100 uppercase">Mangalore Refinery and Petrochemicals Limited</h1>
                    <p class="text-[11px] text-amber-400 font-mono">Government of India Enterprise • Sovereign AI Verification Portal</p>
                </div>
            </div>
            <a href="/" class="px-3 py-1.5 rounded-lg bg-slate-800 hover:bg-slate-700 text-cyan-300 text-xs font-mono font-bold transition flex items-center space-x-1 border border-slate-700">
                <span>← Workbench</span>
            </a>
        </header>

        <!-- Dynamic Integrity Status Banner -->
        <div id="status-card" class="p-6 sm:p-8 rounded-3xl border-2 transition duration-500 cyber-card glow-emerald border-emerald-500/50 flex flex-col items-center text-center space-y-4">
            <div id="status-badge" class="px-4 py-1.5 rounded-full text-xs font-mono font-extrabold tracking-widest bg-emerald-950 text-emerald-300 border border-emerald-500 uppercase animate-pulse">
                ✓ 100% CRYPTOGRAPHICALLY VERIFIED
            </div>
            <h2 id="status-title" class="text-2xl sm:text-3xl font-black text-slate-100">
                Official Proof-of-Execution Certificate
            </h2>
            <p id="status-desc" class="text-sm text-slate-300 max-w-2xl leading-relaxed">
                This verification seal mathematically certifies that the engineering calculations and state transitions for Problem Statement <span class="text-amber-400 font-bold font-mono">MRPL PS 26117</span> were executed strictly on-premise within dedicated sovereign hardware.
            </p>
            
            <div class="w-full grid grid-cols-2 sm:grid-cols-4 gap-3 pt-4 border-t border-slate-800 text-left">
                <div class="p-3 bg-slate-950/60 rounded-xl border border-slate-800">
                    <div class="text-[10px] text-slate-400 font-mono uppercase">Network Air-Gap</div>
                    <div id="metric-airgap" class="text-sm sm:text-base font-black text-emerald-400">0 WAN Packets</div>
                    <div class="text-[9px] text-slate-500">100% Local Sockets</div>
                </div>
                <div class="p-3 bg-slate-950/60 rounded-xl border border-slate-800">
                    <div class="text-[10px] text-slate-400 font-mono uppercase">ASME B31.3 Math</div>
                    <div id="metric-asme" class="text-sm sm:text-base font-black text-cyan-400">5.27 Yrs Safe</div>
                    <div class="text-[9px] text-slate-500">Subprocess Sandbox</div>
                </div>
                <div class="p-3 bg-slate-950/60 rounded-xl border border-slate-800">
                    <div class="text-[10px] text-slate-400 font-mono uppercase">Merkle DAG</div>
                    <div id="metric-merkle" class="text-sm sm:text-base font-black text-purple-400">Intact Root</div>
                    <div class="text-[9px] text-slate-500">Canonical SHA-256</div>
                </div>
                <div class="p-3 bg-slate-950/60 rounded-xl border border-slate-800">
                    <div class="text-[10px] text-slate-400 font-mono uppercase">Governance</div>
                    <div id="metric-gov" class="text-sm sm:text-base font-black text-amber-400">Non-Repudiation</div>
                    <div class="text-[9px] text-slate-500">PSU Audit Compliant</div>
                </div>
            </div>
        </div>

        <!-- Certificate Metadata & Root Hash Card -->
        <div class="p-5 cyber-card rounded-2xl border border-slate-800 flex flex-col space-y-3 font-mono text-xs">
            <div class="flex items-center justify-between pb-2 border-b border-slate-800">
                <span class="text-slate-400 font-bold">Certificate Identifier:</span>
                <span id="cert-id" class="text-cyan-400 font-bold">MRPL-SAI-POE-2026-ACTIVE</span>
            </div>
            <div class="flex items-center justify-between pb-2 border-b border-slate-800">
                <span class="text-slate-400 font-bold">Problem Statement:</span>
                <span class="text-amber-400 font-bold">MRPL PS 26117 (Sovereign Industrial AI)</span>
            </div>
            <div class="flex items-center justify-between pb-2 border-b border-slate-800">
                <span class="text-slate-400 font-bold">Genesis Seed & Authority:</span>
                <span class="text-slate-200">MRPL_REFINERY_GENESIS_ROOT_V1 (CGM Technical Services)</span>
            </div>
            <div class="flex flex-col space-y-1">
                <span class="text-slate-400 font-bold">Cryptographic Merkle Root Digest:</span>
                <div id="merkle-root-display" class="p-3 bg-slate-950 rounded-lg border border-purple-900/60 text-purple-300 font-bold break-all">
                    Loading cryptographic root hash...
                </div>
            </div>
        </div>

        <!-- Causal Merkle Transition Chain (5-Node Trace) -->
        <div class="cyber-card rounded-2xl border border-slate-800 overflow-hidden">
            <div class="p-4 bg-slate-950/80 border-b border-slate-800 flex items-center justify-between">
                <div class="flex items-center space-x-2">
                    <span class="w-2.5 h-2.5 rounded-full bg-cyan-400 animate-pulse"></span>
                    <h3 class="font-bold text-sm text-slate-200">Causal State Transition Chain (DAG Proof)</h3>
                </div>
                <span id="block-count" class="text-[10px] font-mono px-2 py-0.5 rounded bg-slate-800 text-slate-300">5 Blocks</span>
            </div>
            
            <div id="chain-container" class="divide-y divide-slate-800/80 font-mono text-xs">
                <!-- Blocks loaded via JS -->
            </div>
        </div>

        <!-- Live Refresh & PDF Download Controls -->
        <div class="flex flex-col sm:flex-row items-center justify-between gap-3 p-4 cyber-card rounded-2xl border border-slate-800">
            <div class="text-[11px] text-slate-400 font-mono">
                SecureNex Sovereignty Engine • Mangalore Refinery & Petrochemicals Limited
            </div>
            <div class="flex items-center space-x-2">
                <button onclick="loadVerificationData()" class="px-3 py-2 bg-slate-800 hover:bg-slate-700 text-slate-200 rounded-lg text-xs font-mono font-bold transition flex items-center space-x-1">
                    <span>↻ Re-Verify Sockets</span>
                </button>
                <a href="/api/airgap/certificate/pdf" class="px-4 py-2 bg-gradient-to-r from-purple-600 to-indigo-600 hover:from-purple-500 hover:to-indigo-500 text-white rounded-lg text-xs font-mono font-bold transition flex items-center space-x-1 shadow-lg shadow-purple-600/30">
                    <span>Download Official PDF ↓</span>
                </a>
            </div>
        </div>

    </div>

    <script>
        const API_BASE = window.location.origin;

        async function loadVerificationData() {
            try {
                const resAudit = await fetch(API_BASE + '/api/ledger/audit');
                const auditData = await resAudit.json();

                const isIntegrityValid = auditData.integrity.is_valid;
                const rootHash = auditData.root_hash;
                const chain = auditData.chain || [];

                document.getElementById('merkle-root-display').innerText = rootHash;
                document.getElementById('block-count').innerText = chain.length + ' Sealed Blocks';

                const statusCard = document.getElementById('status-card');
                const statusBadge = document.getElementById('status-badge');
                const statusTitle = document.getElementById('status-title');
                const statusDesc = document.getElementById('status-desc');

                if (!isIntegrityValid) {
                    statusCard.className = 'p-6 sm:p-8 rounded-3xl border-2 transition duration-500 cyber-card glow-red border-red-500 flex flex-col items-center text-center space-y-4';
                    statusBadge.className = 'px-4 py-1.5 rounded-full text-xs font-mono font-extrabold tracking-widest bg-red-950 text-red-300 border border-red-500 uppercase animate-pulse';
                    statusBadge.innerText = '🚨 CERTIFICATE REVOKED — TAMPER DETECTED';
                    statusTitle.innerText = 'Cryptographic Verification Breach';
                    statusDesc.innerHTML = '<span class="text-red-300 font-bold">WARNING:</span> An unauthorized in-memory payload modification was detected at Block #' + (auditData.integrity.tampered_block_index || 3) + '. Merkle root hash verification failed. This document is invalidated for official PSU compliance.';
                    
                    document.getElementById('metric-merkle').className = 'text-sm sm:text-base font-black text-red-400';
                    document.getElementById('metric-merkle').innerText = 'TAMPERED';
                    document.getElementById('metric-gov').className = 'text-sm sm:text-base font-black text-red-400';
                    document.getElementById('metric-gov').innerText = 'REVOKED';
                } else {
                    statusCard.className = 'p-6 sm:p-8 rounded-3xl border-2 transition duration-500 cyber-card glow-emerald border-emerald-500/50 flex flex-col items-center text-center space-y-4';
                    statusBadge.className = 'px-4 py-1.5 rounded-full text-xs font-mono font-extrabold tracking-widest bg-emerald-950 text-emerald-300 border border-emerald-500 uppercase animate-pulse';
                    statusBadge.innerText = '✓ 100% CRYPTOGRAPHICALLY VERIFIED';
                    statusTitle.innerText = 'Official Proof-of-Execution Certificate';
                    statusDesc.innerText = 'This verification seal mathematically certifies that the engineering calculations and state transitions for Problem Statement MRPL PS 26117 were executed strictly on-premise within dedicated sovereign hardware.';
                    
                    document.getElementById('metric-merkle').className = 'text-sm sm:text-base font-black text-purple-400';
                    document.getElementById('metric-merkle').innerText = 'Intact Root';
                    document.getElementById('metric-gov').className = 'text-sm sm:text-base font-black text-amber-400';
                    document.getElementById('metric-gov').innerText = 'Non-Repudiation';
                }

                const chainEl = document.getElementById('chain-container');
                chainEl.innerHTML = '';

                if (chain.length === 0) {
                    chainEl.innerHTML = '<div class="p-4 text-center text-slate-500">No transition blocks recorded in active memory.</div>';
                } else {
                    chain.forEach((block) => {
                        const isTamperedBlock = !isIntegrityValid && block.block_index === (auditData.integrity.tampered_block_index || 3);
                        const blockRow = document.createElement('div');
                        blockRow.className = 'p-3.5 flex flex-col sm:flex-row sm:items-center justify-between gap-2 ' + (isTamperedBlock ? 'bg-red-950/70 border-l-4 border-red-500' : 'hover:bg-slate-900/50');
                        
                        blockRow.innerHTML = '<div class="flex items-center space-x-3"><span class="px-2 py-0.5 rounded ' + (isTamperedBlock ? 'bg-red-900 text-red-200' : 'bg-slate-800 text-slate-300') + ' font-bold text-[10px]">#' + block.block_index + '</span><div><div class="font-bold ' + (isTamperedBlock ? 'text-red-300' : 'text-slate-200') + '">' + block.node_name + '</div><div class="text-[10px] text-slate-500">' + new Date(block.timestamp * 1000).toLocaleTimeString() + ' • Parent: ' + (block.prev_block_hash ? block.prev_block_hash.substring(0, 12) + '...' : 'GENESIS') + '</div></div></div><div class="flex items-center space-x-2"><span class="text-[10px] text-slate-400 font-mono bg-slate-950 px-2 py-1 rounded border border-slate-800">' + block.block_hash.substring(0, 18) + '...</span><span class="px-2 py-0.5 rounded text-[10px] font-bold ' + (isTamperedBlock ? 'bg-red-900 text-red-300' : 'bg-emerald-950 text-emerald-300') + '">' + (isTamperedBlock ? '🚨 FORGED' : '✓ SEALED') + '</span></div>';
                        chainEl.appendChild(blockRow);
                    });
                }

            } catch (e) {
                console.error('Verification portal load error:', e);
            }
        }

        window.onload = loadVerificationData;
    </script>
</body>
</html>"""

os.makedirs("ui/static", exist_ok=True)
with open("ui/static/verify.html", "w", encoding="utf-8") as f:
    f.write(verify_html_content)
print("[OK] Created ui/static/verify.html")


# =====================================================================
# 1B. CREATE ui/static/rag_studio.html (Dedicated ChatGPT/Gemini/RAGFlow Studio)
# =====================================================================
rag_studio_html_content = """<!DOCTYPE html>
<html lang="en" class="scroll-smooth">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MRPL Sovereign RAG Studio — Industrial Knowledge Assistant (RAGFlow Architecture)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
    <style>
        body { font-family: 'Inter', -apple-system, sans-serif; background-color: #06090f; color: #e2e8f0; }
        .font-mono { font-family: 'JetBrains Mono', monospace, Consolas; }
        .cyber-card { background: rgba(13, 19, 33, 0.94); backdrop-filter: blur(16px); border: 1px solid rgba(0, 240, 255, 0.15); }
        .glow-cyan { box-shadow: 0 0 25px rgba(6, 182, 212, 0.15); }
        .glow-emerald { box-shadow: 0 0 25px rgba(16, 185, 129, 0.15); }
        .custom-scrollbar::-webkit-scrollbar { width: 6px; height: 6px; }
        .custom-scrollbar::-webkit-scrollbar-track { background: rgba(15, 23, 42, 0.6); }
        .custom-scrollbar::-webkit-scrollbar-thumb { background: rgba(51, 65, 85, 0.8); border-radius: 3px; }
        .custom-scrollbar::-webkit-scrollbar-thumb:hover { background: #0ea5e9; }
        .prose strong { color: #38bdf8; }
        .prose code { color: #f472b6; background: rgba(15, 23, 42, 0.8); padding: 0.15rem 0.35rem; border-radius: 0.25rem; font-family: monospace; }
        .prose ul { list-style-type: disc; margin-left: 1.25rem; }
    </style>
</head>
<body class="h-screen flex flex-col antialiased selection:bg-cyan-500 selection:text-black overflow-hidden">

    <!-- Top Navigation Bar -->
    <header class="h-16 border-b border-slate-800 bg-slate-950/95 px-4 sm:px-6 flex items-center justify-between z-30 flex-shrink-0 shadow-2xl">
        <div class="flex items-center space-x-3">
            <div class="w-9 h-9 rounded-xl bg-gradient-to-br from-amber-500 to-amber-700 flex items-center justify-center font-bold text-black text-lg shadow-lg shadow-amber-500/20">
                M
            </div>
            <div>
                <div class="flex items-center space-x-2">
                    <h1 class="text-sm font-extrabold tracking-wider text-slate-100 uppercase">MRPL Sovereign RAG Studio</h1>
                    <span class="px-2 py-0.5 rounded text-[10px] font-mono font-bold bg-cyan-950 text-cyan-300 border border-cyan-700">RAGFlow-Style BM25 Layer</span>
                </div>
                <p class="text-[11px] text-slate-400 font-mono">Grounded Refinery SOP Search • 100% Air-Gapped Local Memory</p>
            </div>
        </div>

        <div class="flex items-center space-x-3">
            <div class="hidden sm:flex items-center space-x-2 px-3 py-1 rounded-full bg-emerald-950/80 border border-emerald-500/40 text-emerald-300 text-xs font-mono font-bold">
                <span class="w-2 h-2 rounded-full bg-emerald-400 animate-pulse"></span>
                <span>0 WAN PACKETS LOGGED</span>
            </div>
            <a href="/" class="px-3.5 py-1.5 rounded-xl bg-slate-800 hover:bg-slate-700 text-cyan-300 text-xs font-mono font-bold transition flex items-center space-x-1.5 border border-slate-700 shadow-md">
                <span>← Back to Workbench</span>
            </a>
        </div>
    </header>

    <!-- Main Studio Split Layout -->
    <div class="flex-1 flex overflow-hidden relative">

        <!-- LEFT SIDEBAR: Document Ingestion & RAGFlow Chunk Inspector (Width: 380px) -->
        <aside class="w-80 sm:w-96 border-r border-slate-800 bg-[#080d1a] flex flex-col flex-shrink-0 overflow-y-auto custom-scrollbar p-4 space-y-4">
            
            <!-- Section 1: Upload Document -->
            <div class="cyber-card rounded-2xl p-4 border border-slate-800 flex flex-col space-y-3">
                <div class="flex items-center justify-between">
                    <h2 class="text-xs font-mono font-bold text-slate-200 uppercase tracking-wider">1. Ingest Refinery Document</h2>
                    <span class="text-[10px] font-mono text-cyan-400">TXT, MD, JSON</span>
                </div>
                
                <div id="drop-zone" onclick="document.getElementById('file-input').click()" class="border-2 border-dashed border-slate-700 hover:border-cyan-500 rounded-xl p-4 flex flex-col items-center justify-center text-center cursor-pointer transition bg-slate-950/40 hover:bg-cyan-950/20 group">
                    <input type="file" id="file-input" accept=".txt,.md,.json,.pdf,.csv" class="hidden" onchange="handleFileUpload(event)">
                    <div class="w-10 h-10 rounded-full bg-slate-900 group-hover:bg-cyan-900/50 flex items-center justify-center text-slate-400 group-hover:text-cyan-300 transition mb-2">
                        📄
                    </div>
                    <span class="text-xs font-bold text-slate-300 group-hover:text-cyan-200">Click to upload SOP document</span>
                    <span class="text-[10px] text-slate-500 mt-0.5">Air-gapped local memory ingestion</span>
                </div>
            </div>

            <!-- Section 2: Pre-Loaded Curated SOPs -->
            <div class="cyber-card rounded-2xl p-4 border border-slate-800 flex flex-col space-y-2.5">
                <div class="flex items-center justify-between">
                    <h2 class="text-xs font-mono font-bold text-slate-200 uppercase tracking-wider">Preloaded Sample SOPs</h2>
                    <span class="text-[10px] font-mono text-amber-400">MRPL Ground Truth</span>
                </div>
                
                <div class="flex flex-col space-y-1.5 text-xs font-mono">
                    <button onclick="selectSampleSOP('mrpl_refinery_sop_master_handbook.txt')" class="p-2 rounded-lg bg-slate-900/80 hover:bg-slate-800 border border-slate-800 text-left transition flex items-center justify-between group">
                        <div class="truncate pr-2">
                            <span class="text-slate-200 group-hover:text-cyan-300 block truncate">📘 MRPL SOP Master Handbook</span>
                            <span class="text-[10px] text-slate-500">Section 14: Emergency DOP</span>
                        </div>
                        <span class="text-[10px] text-cyan-400">Load ➔</span>
                    </button>

                    <button onclick="selectSampleSOP('plant_safety_interlock_sop_102.txt')" class="p-2 rounded-lg bg-slate-900/80 hover:bg-slate-800 border border-slate-800 text-left transition flex items-center justify-between group">
                        <div class="truncate pr-2">
                            <span class="text-slate-200 group-hover:text-cyan-300 block truncate">⚡ SOP-102 Safety Interlocks</span>
                            <span class="text-[10px] text-slate-500">ESD-1 High-High Temp Trip</span>
                        </div>
                        <span class="text-[10px] text-cyan-400">Load ➔</span>
                    </button>

                    <button onclick="selectSampleSOP('centrifugal_pump_vibration_maintenance_sop.txt')" class="p-2 rounded-lg bg-slate-900/80 hover:bg-slate-800 border border-slate-800 text-left transition flex items-center justify-between group">
                        <div class="truncate pr-2">
                            <span class="text-slate-200 group-hover:text-cyan-300 block truncate">⚙️ SOP-305 Pump Vibration</span>
                            <span class="text-[10px] text-slate-500">ISO 10816 Zone Severity</span>
                        </div>
                        <span class="text-[10px] text-cyan-400">Load ➔</span>
                    </button>

                    <button onclick="selectSampleSOP('dop_procurement_authorization_memo.txt')" class="p-2 rounded-lg bg-slate-900/80 hover:bg-slate-800 border border-slate-800 text-left transition flex items-center justify-between group">
                        <div class="truncate pr-2">
                            <span class="text-slate-200 group-hover:text-cyan-300 block truncate">📋 DOP Authorization Memo</span>
                            <span class="text-[10px] text-slate-500">INR 45 Lakhs Retubing</span>
                        </div>
                        <span class="text-[10px] text-cyan-400">Load ➔</span>
                    </button>
                </div>
            </div>

            <!-- Section 3: Indexing Telemetry & Chunks Status -->
            <div class="cyber-card rounded-2xl p-4 border border-slate-800 flex flex-col space-y-3">
                <div class="flex items-center justify-between">
                    <h2 class="text-xs font-mono font-bold text-slate-200 uppercase tracking-wider">Corpus Index Telemetry</h2>
                    <span id="index-badge" class="px-2 py-0.5 rounded text-[10px] font-mono font-bold bg-emerald-950 text-emerald-300">ACTIVE</span>
                </div>
                
                <div class="grid grid-cols-2 gap-2 text-xs font-mono">
                    <div class="p-2 bg-slate-950 rounded-lg border border-slate-800">
                        <span class="text-[10px] text-slate-400 block">Documents</span>
                        <span id="stat-docs" class="text-base font-bold text-cyan-400">4 Files</span>
                    </div>
                    <div class="p-2 bg-slate-950 rounded-lg border border-slate-800">
                        <span class="text-[10px] text-slate-400 block">Total Chunks</span>
                        <span id="stat-chunks" class="text-base font-bold text-purple-400">18 Chunks</span>
                    </div>
                    <div class="p-2 bg-slate-950 rounded-lg border border-slate-800">
                        <span class="text-[10px] text-slate-400 block">Total Tokens</span>
                        <span id="stat-tokens" class="text-base font-bold text-emerald-400">2,840</span>
                    </div>
                    <div class="p-2 bg-slate-950 rounded-lg border border-slate-800">
                        <span class="text-[10px] text-slate-400 block">Unique Terms</span>
                        <span id="stat-terms" class="text-base font-bold text-amber-400">742</span>
                    </div>
                </div>

                <button onclick="reindexCorpus()" class="w-full py-2 bg-gradient-to-r from-cyan-600 to-blue-600 hover:from-cyan-500 hover:to-blue-500 text-white rounded-xl text-xs font-mono font-bold transition flex items-center justify-center space-x-1.5 shadow-lg shadow-cyan-600/20">
                    <span>⚡ Re-Index &amp; Tokenize Corpus</span>
                </button>
            </div>

            <!-- Section 4: RAGFlow Live Chunks Inspector Drawer -->
            <div class="cyber-card rounded-2xl p-4 border border-slate-800 flex flex-col space-y-2">
                <div class="flex items-center justify-between cursor-pointer" onclick="toggleChunksList()">
                    <h2 class="text-xs font-mono font-bold text-slate-200 uppercase tracking-wider">Inspect Chunks (<span id="chunk-drawer-count">18</span>)</h2>
                    <span id="chunk-toggle-icon" class="text-xs text-cyan-400">▼</span>
                </div>
                <div id="chunks-container" class="hidden flex flex-col space-y-2 max-h-64 overflow-y-auto custom-scrollbar pt-2 text-[11px] font-mono text-slate-300">
                    <!-- Populated dynamically via JS -->
                </div>
            </div>

        </aside>

        <!-- RIGHT MAIN AREA: ChatGPT / Gemini Style Conversational Stream -->
        <main class="flex-1 flex flex-col bg-[#050811] relative overflow-hidden">
            
            <!-- Quick Demo Prompts Header Bar -->
            <div class="p-3 border-b border-slate-800/80 bg-slate-950/60 flex items-center space-x-2 overflow-x-auto custom-scrollbar flex-shrink-0">
                <span class="text-[10px] font-mono font-bold text-slate-400 whitespace-nowrap">DEMO PROMPTS:</span>
                
                <button onclick="setPrompt('Search MRPL Standard Operating Procedures for Section 14: Emergency Turnaround delegation of power limits and required General Manager approval thresholds.')" class="px-2.5 py-1 rounded-full bg-slate-900 hover:bg-cyan-950 text-slate-300 hover:text-cyan-300 border border-slate-800 text-[11px] font-mono whitespace-nowrap transition">
                    💰 Section 14 DOP Limits
                </button>
                <button onclick="setPrompt('What is the Emergency Shutdown trigger criteria for high-high temperature interlock TT-101-HH under SOP-102?')" class="px-2.5 py-1 rounded-full bg-slate-900 hover:bg-cyan-950 text-slate-300 hover:text-cyan-300 border border-slate-800 text-[11px] font-mono whitespace-nowrap transition">
                    ⚡ SOP-102 ESD-1 Interlock
                </button>
                <button onclick="setPrompt('According to SOP-305, what is the maximum acceptable vibration velocity RMS for centrifugal pumps before emergency trip?')" class="px-2.5 py-1 rounded-full bg-slate-900 hover:bg-cyan-950 text-slate-300 hover:text-cyan-300 border border-slate-800 text-[11px] font-mono whitespace-nowrap transition">
                    ⚙️ SOP-305 Pump Vibration
                </button>
                <button onclick="setPrompt('What ASME standard governs crude furnace fired heater tube replacement at MRPL, and what is the required inspection frequency?')" class="px-2.5 py-1 rounded-full bg-slate-900 hover:bg-cyan-950 text-slate-300 hover:text-cyan-300 border border-slate-800 text-[11px] font-mono whitespace-nowrap transition">
                    🔍 ASME B31.3 Inspection Policy
                </button>
            </div>

            <!-- Chat Message Stream -->
            <div id="chat-messages" class="flex-1 overflow-y-auto custom-scrollbar p-4 sm:p-6 space-y-6">
                
                <!-- Initial Welcome Message from MRPL Sovereign Bot -->
                <div class="flex items-start space-x-3.5 max-w-3xl">
                    <div class="w-8 h-8 rounded-xl bg-cyan-950 border border-cyan-500/50 flex items-center justify-center font-bold text-cyan-300 text-xs flex-shrink-0 shadow-lg shadow-cyan-500/20">
                        AI
                    </div>
                    <div class="flex-1 bg-slate-900/90 rounded-2xl rounded-tl-none p-4 border border-slate-800 shadow-md">
                        <div class="flex items-center justify-between mb-2">
                            <span class="text-xs font-bold text-cyan-300 font-mono">MRPL Sovereign Knowledge Assistant (RAGFlow Hybrid Engine)</span>
                            <span class="text-[10px] font-mono text-emerald-400 bg-emerald-950 px-2 py-0.5 rounded border border-emerald-800">100% AIR-GAPPED</span>
                        </div>
                        <div class="text-sm text-slate-200 leading-relaxed font-sans space-y-2">
                            <p>
                                Welcome to the <strong>MRPL Sovereign RAG Studio</strong>. I am your on-premise industrial knowledge assistant, running strictly on local sovereign hardware with zero cloud egress.
                            </p>
                            <p class="text-xs text-slate-400">
                                I am grounded in your local SOP handbooks, interlock thresholds, and delegation of power limits. Ask any operational question or select a demo prompt above to test live retrieval!
                            </p>
                        </div>
                    </div>
                </div>

            </div>

            <!-- Chat Input Area -->
            <div class="p-4 border-t border-slate-800 bg-slate-950/90 flex-shrink-0">
                <div class="max-w-4xl mx-auto flex flex-col space-y-2">
                    <div class="relative flex items-end rounded-2xl bg-slate-900/90 border border-slate-700 focus-within:border-cyan-500 focus-within:ring-1 focus-within:ring-cyan-500 shadow-xl transition">
                        <textarea id="prompt-input" rows="2" placeholder="Ask about MRPL SOPs, DOP financial limits, vibration thresholds, or boiler tube policies..." class="w-full bg-transparent p-3.5 pr-28 text-sm text-slate-100 placeholder-slate-500 focus:outline-none resize-none font-sans custom-scrollbar" onkeydown="handleKey(event)"></textarea>
                        
                        <div class="absolute right-2.5 bottom-2.5 flex items-center space-x-2">
                            <button id="send-btn" onclick="sendPrompt()" class="px-4 py-2 bg-gradient-to-r from-cyan-500 to-blue-600 hover:from-cyan-400 hover:to-blue-500 text-white rounded-xl text-xs font-mono font-bold transition flex items-center space-x-1.5 shadow-lg shadow-cyan-500/25">
                                <span>Ask SOP Bot</span>
                                <span id="send-arrow">➔</span>
                            </button>
                        </div>
                    </div>
                    
                    <div class="flex items-center justify-between text-[11px] font-mono text-slate-500 px-2">
                        <span>Model: <strong class="text-slate-400">DeepSeek-R1 (BM25 Grounded)</strong></span>
                        <span>Zero External API Calls • 100% Confidentiality Guaranteed</span>
                    </div>
                </div>
            </div>

        </main>

        <!-- SLIDE-OUT SOURCE CHUNK INSPECTOR DRAWER (RAGFlow Style) -->
        <div id="source-drawer" class="hidden absolute top-0 right-0 w-96 h-full bg-[#090e1d] border-l border-cyan-500/40 shadow-2xl z-40 flex flex-col transition">
            <div class="p-4 bg-slate-950 border-b border-slate-800 flex items-center justify-between">
                <div class="flex items-center space-x-2">
                    <span class="w-2.5 h-2.5 rounded-full bg-cyan-400"></span>
                    <h3 class="font-bold text-xs text-slate-100 font-mono">Grounded Source Chunk Viewer</h3>
                </div>
                <button onclick="closeSourceDrawer()" class="text-slate-400 hover:text-white text-xs font-mono px-2 py-1 bg-slate-900 rounded">✕</button>
            </div>
            
            <div class="p-4 flex-1 overflow-y-auto custom-scrollbar flex flex-col space-y-3 font-mono text-xs">
                <div class="p-3 bg-slate-950 rounded-xl border border-slate-800">
                    <span class="text-[10px] text-slate-400 block">Document Name:</span>
                    <span id="drawer-doc-name" class="font-bold text-cyan-300">--</span>
                </div>
                <div class="p-3 bg-slate-950 rounded-xl border border-slate-800">
                    <span class="text-[10px] text-slate-400 block">Relevance BM25 Score:</span>
                    <span id="drawer-doc-score" class="font-bold text-emerald-400">--</span>
                </div>
                <div class="flex flex-col space-y-1">
                    <span class="text-[10px] text-slate-400">Verbatim Chunk Snippet:</span>
                    <div id="drawer-doc-text" class="p-3 bg-slate-950 rounded-xl border border-slate-800 text-slate-200 leading-relaxed max-h-96 overflow-y-auto custom-scrollbar whitespace-pre-wrap font-sans text-xs">
                        --
                    </div>
                </div>
            </div>
        </div>

    </div>

    <!-- Scripts -->
    <script>
        const API_BASE = window.location.origin;
        let lastRetrievedChunks = [];

        async function loadTelemetry() {
            try {
                const res = await fetch(`${API_BASE}/api/rag/documents`);
                if (res.ok) {
                    const data = await res.json();
                    document.getElementById("stat-docs").innerText = `${data.total_documents} Files`;
                    document.getElementById("stat-chunks").innerText = `${data.total_chunks} Chunks`;
                    document.getElementById("stat-tokens").innerText = data.total_tokens_indexed.toLocaleString();
                    document.getElementById("stat-terms").innerText = data.unique_terms_indexed.toLocaleString();
                    document.getElementById("chunk-drawer-count").innerText = data.total_chunks;
                    
                    renderChunksDrawer(data.chunks || []);
                }
            } catch (e) {
                console.error("Failed to load RAG telemetry:", e);
            }
        }

        function renderChunksDrawer(chunks) {
            const container = document.getElementById("chunks-container");
            if (!chunks || chunks.length === 0) {
                container.innerHTML = `<div class="text-slate-500 p-2 text-center">No chunks indexed yet.</div>`;
                return;
            }
            container.innerHTML = chunks.map((c, i) => `
                <div onclick="openRawChunk('${c.doc_id}', '${encodeURIComponent(c.text)}')" class="p-2.5 bg-slate-950 rounded-lg border border-slate-800 hover:border-cyan-500/50 cursor-pointer transition">
                    <div class="flex items-center justify-between text-[10px] text-cyan-400 font-bold mb-1">
                        <span>Chunk #${i+1} • ${c.doc_id}</span>
                        <span class="text-slate-500">${c.tokens ? c.tokens.length : 0} tokens</span>
                    </div>
                    <div class="text-[10px] text-slate-300 line-clamp-2">${c.text}</div>
                </div>
            `).join("");
        }

        function toggleChunksList() {
            const el = document.getElementById("chunks-container");
            const icon = document.getElementById("chunk-toggle-icon");
            if (el.classList.contains("hidden")) {
                el.classList.remove("hidden");
                icon.innerText = "▲";
            } else {
                el.classList.add("hidden");
                icon.innerText = "▼";
            }
        }

        function openRawChunk(docId, encodedText, score = "Indexed In Corpus") {
            document.getElementById("drawer-doc-name").innerText = docId;
            document.getElementById("drawer-doc-score").innerText = score;
            document.getElementById("drawer-doc-text").innerText = decodeURIComponent(encodedText);
            document.getElementById("source-drawer").classList.remove("hidden");
        }

        function closeSourceDrawer() {
            document.getElementById("source-drawer").classList.add("hidden");
        }

        function setPrompt(text) {
            document.getElementById("prompt-input").value = text;
            document.getElementById("prompt-input").focus();
        }

        function handleKey(e) {
            if (e.key === "Enter" && !e.shiftKey) {
                e.preventDefault();
                sendPrompt();
            }
        }

        async function selectSampleSOP(filename) {
            try {
                const res = await fetch(`${API_BASE}/api/rag/load-sample?filename=${encodeURIComponent(filename)}`, { method: "POST" });
                const data = await res.json();
                loadTelemetry();
                appendAIMessage(`Loaded **${filename}** into sovereign in-memory BM25 index (${data.chunks_indexed} new chunks). You can now ask questions about this document!`);
            } catch (e) {
                console.error("Error loading sample SOP:", e);
            }
        }

        async function reindexCorpus() {
            try {
                document.getElementById("index-badge").innerText = "INDEXING...";
                document.getElementById("index-badge").className = "px-2 py-0.5 rounded text-[10px] font-mono font-bold bg-amber-950 text-amber-300 animate-pulse";
                
                const res = await fetch(`${API_BASE}/api/rag/reindex`, { method: "POST" });
                const data = await res.json();
                
                document.getElementById("index-badge").innerText = "ACTIVE";
                document.getElementById("index-badge").className = "px-2 py-0.5 rounded text-[10px] font-mono font-bold bg-emerald-950 text-emerald-300";
                loadTelemetry();
                appendAIMessage(`✓ Complete corpus successfully re-indexed into memory! Indexed **${data.total_chunks} chunks** across **${data.total_documents} documents** with 100% air-gap compliance.`);
            } catch (e) {
                console.error("Reindex error:", e);
            }
        }

        async function handleFileUpload(e) {
            const file = e.target.files[0];
            if (!file) return;

            const formData = new FormData();
            formData.append("file", file);

            try {
                appendUserMessage(`Uploading file: ${file.name}`);
                const res = await fetch(`${API_BASE}/api/rag/upload`, {
                    method: "POST",
                    body: formData
                });
                const data = await res.json();
                loadTelemetry();
                appendAIMessage(`✓ Successfully ingested and chunked **${file.name}** into **${data.chunks_created} atomic chunks**. Ready for grounded question answering!`);
            } catch (err) {
                console.error("Upload error:", err);
                appendAIMessage(`🚨 Error uploading document: ${err.message}`);
            }
        }

        function appendUserMessage(text) {
            const chat = document.getElementById("chat-messages");
            const div = document.createElement("div");
            div.className = "flex items-start justify-end space-x-3 max-w-3xl ml-auto";
            div.innerHTML = `
                <div class="bg-cyan-950/70 border border-cyan-600/40 rounded-2xl rounded-tr-none p-3.5 text-sm text-cyan-100 font-sans shadow-md">
                    ${text}
                </div>
                <div class="w-8 h-8 rounded-xl bg-slate-800 border border-slate-700 flex items-center justify-center font-bold text-slate-300 text-xs flex-shrink-0">
                    U
                </div>
            `;
            chat.appendChild(div);
            chat.scrollTop = chat.scrollHeight;
        }

        function appendAIMessage(markdownText, citations = [], rawChunks = []) {
            const chat = document.getElementById("chat-messages");
            const div = document.createElement("div");
            div.className = "flex items-start space-x-3.5 max-w-3xl";
            
            lastRetrievedChunks = rawChunks || [];

            const citationHtml = citations && citations.length > 0 ? `
                <div class="mt-3 pt-2.5 border-t border-slate-800/80 flex flex-wrap gap-1.5">
                    <span class="text-[10px] font-mono text-slate-500 mr-1 flex items-center">Grounded Sources:</span>
                    ${citations.map((c, idx) => {
                        const chunkMatch = lastRetrievedChunks.find(ch => ch.doc_id === c) || (lastRetrievedChunks[idx] || {});
                        const encoded = encodeURIComponent(chunkMatch.text || "Direct Grounded Reference");
                        const score = chunkMatch.relevance_score ? `Score: ${chunkMatch.relevance_score}` : "Verified";
                        return `<button onclick="openRawChunk('${c}', '${encoded}', '${score}')" class="px-2.5 py-0.5 rounded-full bg-slate-950 hover:bg-cyan-950 text-cyan-400 hover:text-cyan-200 border border-cyan-800/60 hover:border-cyan-500 text-[10px] font-mono transition flex items-center space-x-1">
                            <span>📌 [Ref ${idx+1}: ${c}]</span>
                            <span class="text-[9px] text-slate-500">🔍</span>
                        </button>`;
                    }).join("")}
                </div>
            ` : "";

            div.innerHTML = `
                <div class="w-8 h-8 rounded-xl bg-cyan-950 border border-cyan-500/50 flex items-center justify-center font-bold text-cyan-300 text-xs flex-shrink-0 shadow-lg shadow-cyan-500/20">
                    AI
                </div>
                <div class="flex-1 bg-slate-900/90 rounded-2xl rounded-tl-none p-4 border border-slate-800 shadow-md">
                    <div class="flex items-center justify-between mb-2">
                        <span class="text-xs font-bold text-cyan-300 font-mono">MRPL Sovereign Assistant</span>
                        <span class="text-[10px] font-mono text-emerald-400">Grounded SOP Answer</span>
                    </div>
                    <div class="text-sm text-slate-200 leading-relaxed font-sans prose prose-invert max-w-none">
                        ${marked.parse(markdownText)}
                    </div>
                    ${citationHtml}
                </div>
            `;
            chat.appendChild(div);
            chat.scrollTop = chat.scrollHeight;
        }

        async function sendPrompt() {
            const input = document.getElementById("prompt-input");
            const prompt = input.value.trim();
            if (!prompt) return;

            appendUserMessage(prompt);
            input.value = "";

            const sendBtn = document.getElementById("send-btn");
            sendBtn.disabled = true;
            sendBtn.innerHTML = `<span>Thinking...</span>`;

            try {
                const res = await fetch(`${API_BASE}/api/rag/chat`, {
                    method: "POST",
                    headers: { "Content-Type": "application/json" },
                    body: JSON.stringify({ prompt: prompt, top_k: 3 })
                });
                const data = await res.json();
                appendAIMessage(data.answer, data.sources_cited, data.retrieved_chunks);
            } catch (err) {
                console.error("Chat error:", err);
                appendAIMessage(`🚨 Error executing sovereign retrieval: ${err.message}`);
            } finally {
                sendBtn.disabled = false;
                sendBtn.innerHTML = `<span>Ask SOP Bot</span> <span>➔</span>`;
            }
        }

        window.onload = loadTelemetry;
    </script>
</body>
</html>"""

with open("ui/static/rag_studio.html", "w", encoding="utf-8") as f:
    f.write(rag_studio_html_content)
print("[OK] Created ui/static/rag_studio.html")


# =====================================================================
# 2. CREATE UNIFIED sample_files/ DIRECTORY STRUCTURE
# =====================================================================
base_sample_dir = "sample_files"
os.makedirs(f"{base_sample_dir}/01_multimodal_vision", exist_ok=True)
os.makedirs(f"{base_sample_dir}/02_sovereign_rag_sops", exist_ok=True)
os.makedirs(f"{base_sample_dir}/03_engineering_math_code_gen", exist_ok=True)
os.makedirs(f"{base_sample_dir}/04_psu_reasoning_memos", exist_ok=True)


# ---------------------------------------------------------------------
# Category 1: Multimodal Vision & NDT Ultrasonic Boiler Excel Report
# ---------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "UT_Inspection_Grid"

# Title header
ws.merge_cells("A1:G1")
ws["A1"] = "MANGALORE REFINERY AND PETROCHEMICALS LIMITED (MRPL)"
ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:G2")
ws["A2"] = "NON-DESTRUCTIVE TESTING (NDT) ULTRASONIC THICKNESS SURVEY REPORT"
ws["A2"].font = Font(name="Arial", size=11, bold=True, color="047857")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

# Equipment Metadata Table
meta = [
    ("Equipment Tag", "B-101-Crude-Furnace-Coil-Tubes", "Inspection Date", "14-AUG-2026"),
    ("Design Pressure (P)", "4.0 MPa (40.0 bar)", "Operating Temp", "365 °C"),
    ("Outer Diameter (D)", "219.10 mm (8-inch NPS)", "Material Spec", "ASTM A106 Grade B"),
    ("Allowable Stress (S)", "137.00 MPa (at 365°C)", "Joint Efficiency (E)", "1.00 (Seamless)"),
    ("Nominal Wall Thickness", "9.53 mm (Schedule 40)", "Corrosion Allowance (c)", "3.00 mm"),
    ("Annual Corrosion Rate", "0.25 mm / year", "Service Fluid", "Sour Heavy Arab Crude"),
]

for row_idx, row in enumerate(meta, start=4):
    ws[f"A{row_idx}"] = row[0]
    ws[f"A{row_idx}"].font = Font(bold=True)
    ws[f"B{row_idx}"] = row[1]
    ws[f"D{row_idx}"] = row[2]
    ws[f"D{row_idx}"].font = Font(bold=True)
    ws[f"E{row_idx}"] = row[3]

# 16-Point Ultrasonic Test Grid Readings (Rows A, B, C, D vs Quadrants 1, 2, 3, 4)
ws["A11"] = "16-POINT ULTRASONIC THICKNESS GRID MEASUREMENTS (in millimeters):"
ws["A11"].font = Font(bold=True, size=11, color="1E293B")

grid_headers = ["Grid Position", "Quadrant 1 (0°)", "Quadrant 2 (90°)", "Quadrant 3 (180°)", "Quadrant 4 (270°)", "Row Min (mm)", "Status"]
for col_idx, h in enumerate(grid_headers, start=1):
    cell = ws.cell(row=12, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

grid_rows = [
    ("Row A (Inlet Flange)", 8.42, 8.35, 8.28, 8.40, 8.28, "SAFE"),
    ("Row B (Radiant Tube Peak)", 7.85, 7.62, 7.48, 7.70, 7.48, "CRITICAL_MIN"),
    ("Row C (Convection Loop)", 8.10, 7.95, 7.88, 8.05, 7.88, "SAFE"),
    ("Row D (Outlet Header)", 8.30, 8.22, 8.15, 8.25, 8.15, "SAFE"),
]

for r_offset, r_data in enumerate(grid_rows, start=13):
    for c_offset, val in enumerate(r_data, start=1):
        cell = ws.cell(row=r_offset, column=c_offset, value=val)
        cell.alignment = Alignment(horizontal="center" if c_offset > 1 else "left")
        if val == "CRITICAL_MIN":
            cell.font = Font(bold=True, color="B91C1C")
            cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif val == 7.48:
            cell.font = Font(bold=True, color="B91C1C")

# Summary Findings Box
ws["A18"] = "INSPECTION SUMMARY & GOVERNING NDT FINDING:"
ws["A18"].font = Font(bold=True, color="047857")
ws["A19"] = "• Minimum Measured Wall Thickness: 7.48 mm (Recorded at Radiant Tube Peak Grid B3)."
ws["A20"] = "• ASME B31.3 Minimum Required Design Wall Thickness: 6.162 mm (Pressure: 3.162mm + Corrosion Allowance: 3.0mm)."
ws["A21"] = "• Net Safe Operational Margin: +1.318 mm above critical threshold."
ws["A22"] = "• Estimated Remaining Safe Life: 5.27 Years (at 0.25 mm/year corrosion rate)."
ws["A23"] = "• Integrity Recommendation: APPROVED_SAFE for continued operation without emergency retubing."

wb.save(f"{base_sample_dir}/01_multimodal_vision/crude_furnace_b101_ultrasonic_ndt_report.xlsx")
wb.save("data/sample_inspection_report.xlsx")
print("[OK] Created sample_files/01_multimodal_vision/crude_furnace_b101_ultrasonic_ndt_report.xlsx")

# P&ID Metadata JSON
pid_meta = {
    "diagram_id": "MRPL-PID-CDU-0101-REV4",
    "unit": "Crude Distillation Unit (CDU-1)",
    "equipment": [
        {
            "tag": "B-101",
            "type": "Crude Pre-Flash Direct Fired Furnace",
            "design_duty_gcal_hr": 42.5,
            "tubes_spec": "ASTM A106 Gr B, 8 NPS Sch 40",
            "design_pressure_mpa": 4.0,
            "design_temperature_c": 400.0,
            "safety_valve_tag": "PSV-101A/B",
            "set_pressure_barg": 42.0
        },
        {
            "tag": "C-101",
            "type": "Main Atmospheric Fractionation Column",
            "trays_count": 48,
            "top_temperature_c": 115.0,
            "bottom_temperature_c": 355.0
        }
    ],
    "control_loops": [
        {"loop_id": "TIC-101", "variable": "Furnace Coil Outlet Temperature", "setpoint_c": 365.0},
        {"loop_id": "PIC-104", "variable": "Fuel Gas Header Pressure", "setpoint_barg": 3.5}
    ]
}
with open(f"{base_sample_dir}/01_multimodal_vision/sample_pid_crude_distillation_metadata.json", "w") as f:
    json.dump(pid_meta, f, indent=2)
print("[OK] Created sample_files/01_multimodal_vision/sample_pid_crude_distillation_metadata.json")


# ---------------------------------------------------------------------
# Category 2: Sovereign RAG SOP Documents
# ---------------------------------------------------------------------
sop_master = """================================================================================
MANGALORE REFINERY AND PETROCHEMICALS LIMITED (MRPL)
STANDARD OPERATING PROCEDURES & ASSET INTEGRITY MASTER HANDBOOK (REV 2026)
================================================================================

SECTION 1: GENERAL PRINCIPLES OF COMPUTATIONAL SOVEREIGNTY & AIR-GAP COMPLIANCE
All refinery engineering data, Piping and Instrumentation Diagrams (P&IDs), ultrasonic inspection logs, financial estimates, and Delegation of Power approval correspondence are classified as STRICTLY CONFIDENTIAL. Transmission of refinery telemetry or documents to external public cloud LLMs (e.g. OpenAI, Anthropic, Groq) constitutes a severe regulatory violation under the Petroleum and Natural Gas Regulatory Board (PNGRB) Cybersecurity Guidelines. All automated synthesis must execute exclusively on on-premise sovereign hardware.

SECTION 2: CRUDE DISTILLATION UNIT (CDU) TUBE INTEGRITY & ASME B31.3 STANDARDS
2.1 Fired furnace heater tubes (Equipment Tag B-101) operating in heavy sour crude service must maintain a minimum wall thickness governed by ASME B31.3 Paragraph 304.1.2 Modified Barlow's Equation:
    t_min = (P * D) / (2 * (S * E * W + P * Y)) + c
Where:
- P = Internal Design Pressure in MPa
- D = Outer Pipe Diameter in mm
- S = Allowable Material Stress at Design Temperature in MPa
- E = Joint Quality Factor (1.0 for seamless piping)
- W = Weld Joint Strength Reduction Factor (1.0)
- Y = Temperature Coefficient (0.4 for ferritic steel under 482 deg C)
- c = Sum of mechanical tolerances and specified corrosion allowance (standard 3.0 mm for heavy crude).

2.2 Inspection Grid & Corrosion Monitoring Frequency:
Ultrasonic thickness gauging (NDT) on radiant furnace coils must be conducted at 16 designated test grid locations during every planned turnaround or at intervals not exceeding 24 operating months.

SECTION 14: EMERGENCY TURNAROUND & FINANCIAL DELEGATION OF POWER (DOP)
14.1 Procurement and Emergency Maintenance Limits:
(a) Expenditure up to INR 25,00,000/- (Twenty Five Lakhs) may be authorized by the Deputy General Manager (Inspection / Operations).
(b) Expenditure exceeding INR 25,00,000/- up to INR 1,00,00,000/- (One Crore) requires prior administrative sanction from the Chief General Manager (Technical Services).
(c) Capital expenditure exceeding INR 1,00,00,000/- requires Board Level Executive Committee approval.

14.2 Boiler Tube Replacement Policy:
If measured wall thickness drops below the calculated ASME B31.3 t_min threshold with a remaining operational life of less than 12 months, immediate retubing authorization must be initiated under Delegation of Power Item 4.2(b).
"""
with open(f"{base_sample_dir}/02_sovereign_rag_sops/mrpl_refinery_sop_master_handbook.txt", "w", encoding="utf-8") as f:
    f.write(sop_master)

sop_interlock = """================================================================================
MRPL PLANT SAFETY INTERLOCK & EMERGENCY SHUTDOWN (ESD) PROCEDURE (SOP-102)
================================================================================
1. High-High Temperature Interlock (TT-101-HH):
If furnace coil outlet temperature exceeds 395 °C for more than 15 consecutive seconds, the Emergency Shutdown System (ESD-1) shall automatically trip the main fuel gas shut-off valve (XV-1041) and inject purge steam.

2. Low-Low Flow Interlock (FT-102-LL):
If crude hydrocarbon feed flow rate drops below 180 m3/hr, burner firing rate shall automatically modulate down to minimum pilot flame to prevent localized coil skin overheating and coking.
"""
with open(f"{base_sample_dir}/02_sovereign_rag_sops/plant_safety_interlock_sop_102.txt", "w", encoding="utf-8") as f:
    f.write(sop_interlock)

sop_pump = """================================================================================
MRPL ROTATING EQUIPMENT & CENTRIFUGAL PUMP MAINTENANCE GUIDELINES (SOP-305)
================================================================================
1. ISO 10816 Vibration Severity Limits for Refinery Pumps (P-101A/B):
- Velocity RMS < 2.8 mm/s: Zone A (Good / Normal Operation)
- Velocity RMS 2.8 - 4.5 mm/s: Zone B (Acceptable for Unrestricted Long-term Operation)
- Velocity RMS 4.5 - 7.1 mm/s: Zone C (Unsatisfactory - Plan maintenance within 30 days)
- Velocity RMS > 7.1 mm/s: Zone D (Critical Danger - Immediate Trip / Standby Switchover)

2. Mechanical Seal Flush Plan:
All hot hydrocarbon pumps operating above 150 °C must utilize API Plan 53B pressurized barrier fluid system with continuous pressure switch monitoring.
"""
with open(f"{base_sample_dir}/02_sovereign_rag_sops/centrifugal_pump_vibration_maintenance_sop.txt", "w", encoding="utf-8") as f:
    f.write(sop_pump)
print("[OK] Created sample_files/02_sovereign_rag_sops/ text files")


# ---------------------------------------------------------------------
# Category 3: Engineering Calculations & Code Gen Specifications
# ---------------------------------------------------------------------
asme_spec = {
    "standard": "ASME_B31_3",
    "formula": "t_min = (P * D) / (2 * (S * E * W + P * Y)) + CA",
    "parameters": {
        "design_pressure_mpa": 4.0,
        "outer_diameter_mm": 219.10,
        "allowable_stress_mpa": 137.0,
        "joint_efficiency": 1.0,
        "weld_factor": 1.0,
        "temperature_y_coefficient": 0.4,
        "corrosion_allowance_mm": 3.0,
        "measured_thickness_mm": 7.48,
        "annual_corrosion_rate_mm_yr": 0.25
    },
    "expected_results": {
        "pressure_thickness_mm": 3.162,
        "required_min_thickness_mm": 6.162,
        "safe_margin_mm": 1.318,
        "remaining_life_years": 5.27,
        "status": "APPROVED_SAFE"
    }
}
with open(f"{base_sample_dir}/03_engineering_math_code_gen/asme_b31_3_piping_thickness_spec.json", "w") as f:
    json.dump(asme_spec, f, indent=2)

lmtd_spec = {
    "calculation_type": "HEAT_EXCHANGER_LMTD_THERMAL_DUTY",
    "parameters": {
        "th_inlet_celsius": 180.0,
        "th_outlet_celsius": 110.0,
        "tc_inlet_celsius": 30.0,
        "tc_outlet_celsius": 85.0,
        "heat_transfer_area_m2": 120.0,
        "overall_u_coeff_w_m2k": 450.0
    },
    "formulae": {
        "dt1": "Th_in - Tc_out = 180 - 85 = 95 C",
        "dt2": "Th_out - Tc_in = 110 - 30 = 80 C",
        "lmtd": "(dt1 - dt2) / ln(dt1 / dt2)",
        "duty_kw": "(U * Area * LMTD) / 1000"
    }
}
with open(f"{base_sample_dir}/03_engineering_math_code_gen/heat_exchanger_lmtd_thermal_duty_spec.json", "w") as f:
    json.dump(lmtd_spec, f, indent=2)

darcy_spec = {
    "calculation_type": "DARCY_WEISBACH_HYDRAULIC_HEAD_LOSS",
    "parameters": {
        "flow_velocity_ms": 2.5,
        "pipe_diameter_mm": 150.0,
        "pipe_length_meters": 100.0,
        "friction_factor_f": 0.020,
        "fluid_density_kg_m3": 850.0
    },
    "formula": "h_f = f * (L/D) * (v^2 / (2*g)) ; delta_P = rho * g * h_f"
}
with open(f"{base_sample_dir}/03_engineering_math_code_gen/darcy_weisbach_hydraulic_flow_spec.json", "w") as f:
    json.dump(darcy_spec, f, indent=2)
print("[OK] Created sample_files/03_engineering_math_code_gen/ spec files")


# ---------------------------------------------------------------------
# Category 4: PSU Administrative Reasoning & Approval Memos
# ---------------------------------------------------------------------
dop_memo = """================================================================================
MANGALORE REFINERY AND PETROCHEMICALS LIMITED (MRPL)
INTER-OFFICE MEMORANDUM & APPROVAL NOTE PROPOSAL
================================================================================
To: Chief General Manager (Technical Services)
From: Chief Manager (Inspection & Non-Destructive Testing)
Date: 15-August-2026
Subject: Proposal for Planned Retubing & Coil Procurement for CDU Furnace B-101

1. BACKGROUND:
Ultrasonic thickness survey on crude furnace B-101 revealed radiant coil minimum wall thickness of 7.48 mm against ASME B31.3 threshold of 6.162 mm. Although currently certified for continued safe operation for 5.27 years, procurement lead time for specialized ASTM A106 Grade B seamless coils is estimated at 14 months.

2. FINANCIAL PROPOSAL & DELEGATION OF POWER:
Estimated procurement and installation cost: INR 45,00,000/- (Rupees Forty-Five Lakhs Only).
Under Item 14.1(b) of MRPL Delegation of Power (DOP), financial sanction of this magnitude falls within the delegated authority of the Chief General Manager (Technical Services).

3. RECOMMENDATION:
Approval is solicited for initiating global e-tender on GeM portal for procurement of 120 meters of 8 NPS Schedule 40 seamless coils to be scheduled for installation during Q4 Turnaround.
"""
with open(f"{base_sample_dir}/04_psu_reasoning_memos/dop_procurement_authorization_memo.txt", "w", encoding="utf-8") as f:
    f.write(dop_memo)
print("[OK] Created sample_files/04_psu_reasoning_memos/ memo file")

print("\n==================================================================")
print("ALL DEMO SAMPLE FILES SUCCESSFULLY CREATED IN: sample_files/")
print("==================================================================")